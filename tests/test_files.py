import csv
import os.path
import time
import zipfile
import requests
import xlrd
from selenium import webdriver
from selene import browser
from pypdf import PdfReader
from openpyxl import load_workbook

CURRENT_FILE_PATH = os.path.abspath(__file__)
DIR_PATH = os.path.dirname(CURRENT_FILE_PATH)
RESOURCES_PATH = os.path.join(DIR_PATH, '..', 'resources', )
DOWNLOAD_PATH = os.path.join(DIR_PATH, 'download')


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_csv():
    csv_path = os.path.join(RESOURCES_PATH, 'eggs.csv')
    with open(csv_path, 'w') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',')
        csvwriter.writerow(['Anna', 'Pavel', 'Peter'])
        csvwriter.writerow(['Alex', 'Serj', 'Yana'])

    with open(csv_path) as csvfile:
        csvreader = csv.reader(csvfile)
        name = []
        for row in csvreader:
            name.append(row)
        assert name[0] == ['Anna', 'Pavel', 'Peter']
        assert name[2] == ['Alex', 'Serj', 'Yana']


# TODO оформить в тест, добавить ассерты и использовать универсальный путь к tmp
def test_download_file_with_browser():
    DOWNLOAD_FILE_PATH = os.path.join(DOWNLOAD_PATH, 'pytest-main.zip')
    if not os.path.exists(DOWNLOAD_PATH):
        os.mkdir(DOWNLOAD_PATH)
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": DOWNLOAD_PATH,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)
    browser.config.driver_options = options
    browser.open("https://github.com/pytest-dev/pytest")
    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()
    time.sleep(1)
    size = os.path.getsize(DOWNLOAD_FILE_PATH)
    assert size == 1564360
    os.remove(DOWNLOAD_FILE_PATH)


def test_download_file_with_requests():
    # TODO сохранять и читать из tmp, использовать универсальный путь
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'
    if not os.path.exists(DOWNLOAD_PATH):
        os.mkdir(DOWNLOAD_PATH)
    PNG_PATH = os.path.join(DOWNLOAD_PATH, 'selenium_logo.png')
    r = requests.get(url)
    with open(PNG_PATH, 'wb') as file:
        file.write(r.content)
    size = os.path.getsize(PNG_PATH)
    assert size == 30803
    os.remove(PNG_PATH)


def test_pdf():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    PDF_PATH = os.path.join(RESOURCES_PATH, 'docs-pytest-org-en-latest.pdf')
    reader = PdfReader(PDF_PATH)
    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()
    print('\n', page)
    print(number_of_pages)
    print(text)
    assert number_of_pages == 412
    assert text == 'pytest Documentation\nRelease 0.1\nholger krekel, trainer and consultant, https://merlinux.eu/\nJul 14, 2022'


def test_xls():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    XLS_PATH = os.path.join(RESOURCES_PATH, 'file_example_XLS_10.xls')
    book = xlrd.open_workbook(XLS_PATH)
    print(f'Количество листов {book.nsheets}')
    print(f'Имена листов {book.sheet_names()}')
    sheet = book.sheet_by_index(0)
    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')
    # печать всех строк по очереди
    for rx in range(sheet.nrows):
        print(sheet.row(rx))
    assert book.nsheets == 1
    assert book.sheet_names() == ['Sheet1']
    assert sheet.ncols == 8
    assert sheet.nrows == 10
    assert sheet.cell_value(rowx=0, colx=1) == 'First Name'


def test_xlsx():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    XLSX_PATH = os.path.join(RESOURCES_PATH, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(XLSX_PATH)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'
    assert sheet.cell(row=3, column=3).value == 'Hashimoto'
    assert sheet.cell(row=3, column=8).value == 1582


def test_add_files_to_zip():
    zip_name = "test.zip"
    with zipfile.ZipFile(zip_name, "w") as zip_file:
        for file_name in os.listdir(RESOURCES_PATH):
            file_path = os.path.join(RESOURCES_PATH, file_name)
            zip_file.write(file_path, file_name)

    with zipfile.ZipFile(zip_name, "r") as zip_file:
        for file_name in os.listdir(RESOURCES_PATH):
            assert file_name in zip_file.namelist()

    os.remove(zip_name)
